from openpyxl import load_workbook

def parse(filepath):
    try:
        # Load the Excel workbook
        workbook = load_workbook(filepath, data_only=True)
        content = ""
        
        # Process all worksheets
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            content += f"=== Sheet: {sheet_name} ===\n"
            
            # Extract data from all cells
            for row in sheet.iter_rows(values_only=True):
                row_text = []
                for cell_value in row:
                    if cell_value is not None:
                        row_text.append(str(cell_value))
                if row_text:  # Only add non-empty rows
                    content += "\t".join(row_text) + "\n"
            content += "\n"
        
        return {
            "status": "success",
            "filename": filepath,
            "content": content.strip(),
            "error": None
        }
    except Exception as e:
        return {
            "status": "error",
            "filename": filepath,
            "content": None,
            "error": str(e)
        }